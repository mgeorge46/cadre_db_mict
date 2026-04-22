from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db.models import Q, Count
from django.utils import timezone
import io
import json

from employees.models import Employee, EmploymentHistory
from core.models import CadreCategory, Position, JobRank, Ministry, Agency, GovernmentDepartment, District, EmployeeType


def admin_required(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('accounts:login')
        if not (request.user.is_admin or request.user.is_it_admin or request.user.is_superuser):
            messages.error(request, 'Access denied.')
            return redirect('dashboard:index')
        return view_func(request, *args, **kwargs)
    wrapper.__name__ = view_func.__name__
    return wrapper


@admin_required
def reports_index(request):
    context = {
        'page_title': 'Reports',
        'breadcrumbs': [('Reports', None)],
        'cadre_categories': CadreCategory.objects.filter(is_active=True),
        'positions': Position.objects.filter(is_active=True),
        'job_ranks': JobRank.objects.filter(is_active=True),
        'ministries': Ministry.objects.filter(is_active=True),
        'agencies': Agency.objects.filter(is_active=True),
        'departments': GovernmentDepartment.objects.filter(is_active=True),
        'districts': District.objects.filter(is_active=True),
        'employee_types': EmployeeType.objects.filter(is_active=True),
    }
    return render(request, 'reports/reports.html', context)


@admin_required
def reports_preview(request):
    from django.core.paginator import Paginator
    report_type = request.GET.get('report_type', 'employee_list')
    employees = get_filtered_employees(request)

    per_page = request.GET.get('per_page', '50')
    try:
        per_page = int(per_page)
        if per_page not in (20, 50, 100):
            per_page = 50
    except (ValueError, TypeError):
        per_page = 50

    paginator = Paginator(employees, per_page)
    page_obj = paginator.get_page(request.GET.get('page'))

    total = employees.count()

    # Summary stats for charts
    cadre_dist = list(employees.values('cadre_category__name').annotate(count=Count('id')).order_by('-count'))
    entity_dist = list(employees.values('entity_type').annotate(count=Count('id')).order_by('-count'))
    gender_dist = list(employees.values('gender').annotate(count=Count('id')).order_by('-count'))

    # Entity type display names
    entity_map = dict(Employee._meta.get_field('entity_type').choices)
    for e in entity_dist:
        e['entity_type'] = entity_map.get(e['entity_type'], e['entity_type'] or 'Unknown')

    for g in gender_dist:
        g['gender'] = {'M': 'Male', 'F': 'Female', '': 'Not Set'}.get(g['gender'], g['gender'] or 'Not Set')

    for c in cadre_dist:
        c['cadre_category__name'] = c['cadre_category__name'] or 'Unassigned'

    context = {
        'page_title': 'Report Preview',
        'breadcrumbs': [('Reports', 'reports:index'), ('Preview', None)],
        'page_obj': page_obj,
        'per_page': per_page,
        'report_type': report_type,
        'total_count': total,
        'cadre_dist_json': json.dumps(cadre_dist),
        'entity_dist_json': json.dumps(entity_dist),
        'gender_dist_json': json.dumps(gender_dist),
        'cadre_categories': CadreCategory.objects.filter(is_active=True),
        'positions': Position.objects.filter(is_active=True),
        'job_ranks': JobRank.objects.filter(is_active=True),
        'ministries': Ministry.objects.filter(is_active=True),
        'agencies': Agency.objects.filter(is_active=True),
        'departments': GovernmentDepartment.objects.filter(is_active=True),
        'districts': District.objects.filter(is_active=True),
        'employee_types': EmployeeType.objects.filter(is_active=True),
        'query_string': request.GET.urlencode(),
    }
    return render(request, 'reports/reports_preview.html', context)


def get_filtered_employees(request):
    from datetime import date, timedelta
    qs = Employee.objects.select_related(
        'user', 'cadre_category', 'position', 'job_rank', 'employee_type',
        'ministry', 'agency', 'government_department', 'district'
    ).filter(is_active=True)

    entity_type = request.GET.get('entity_type', '')
    cadre_cat = request.GET.get('cadre_category', '')
    position_id = request.GET.get('position', '')
    job_rank_id = request.GET.get('job_rank', '')
    emp_type = request.GET.get('employee_type', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    # Time in position filter
    years_operator = request.GET.get('years_operator', '')
    years_value = request.GET.get('years_value', '')
    # Age filter
    age_operator = request.GET.get('age_operator', '')
    age_value = request.GET.get('age_value', '')
    # Inquiry status filter
    inquiry_status = request.GET.get('inquiry_status', '')

    if entity_type:
        qs = qs.filter(entity_type=entity_type)
    if cadre_cat:
        qs = qs.filter(cadre_category_id=cadre_cat)
    if position_id:
        qs = qs.filter(position_id=position_id)
    if job_rank_id:
        qs = qs.filter(job_rank_id=job_rank_id)
    if emp_type:
        qs = qs.filter(employee_type_id=emp_type)
    if date_from:
        qs = qs.filter(date_joined_ministry__gte=date_from)
    if date_to:
        qs = qs.filter(date_joined_ministry__lte=date_to)

    # Time in position filter (years)
    if years_value and years_operator:
        try:
            yval = float(years_value)
            today = date.today()
            cutoff = today - timedelta(days=int(yval * 365.25))
            if years_operator == 'gt':
                qs = qs.filter(date_joined_position__lte=cutoff, date_joined_position__isnull=False)
            elif years_operator == 'gte':
                qs = qs.filter(date_joined_position__lte=cutoff, date_joined_position__isnull=False)
            elif years_operator == 'lt':
                qs = qs.filter(date_joined_position__gte=cutoff, date_joined_position__isnull=False)
            elif years_operator == 'lte':
                qs = qs.filter(date_joined_position__gte=cutoff, date_joined_position__isnull=False)
            elif years_operator == 'eq':
                lower = today - timedelta(days=int((yval + 0.5) * 365.25))
                upper = today - timedelta(days=int((yval - 0.5) * 365.25))
                qs = qs.filter(date_joined_position__gte=lower, date_joined_position__lte=upper, date_joined_position__isnull=False)
        except (ValueError, TypeError):
            pass

    # Age filter
    if age_value and age_operator:
        try:
            aval = int(age_value)
            today = date.today()
            try:
                cutoff_dob = date(today.year - aval, today.month, today.day)
            except ValueError:
                cutoff_dob = date(today.year - aval, today.month, 28)
            if age_operator == 'gt':
                qs = qs.filter(date_of_birth__lte=cutoff_dob, date_of_birth__isnull=False)
            elif age_operator == 'gte':
                qs = qs.filter(date_of_birth__lte=cutoff_dob, date_of_birth__isnull=False)
            elif age_operator == 'lt':
                qs = qs.filter(date_of_birth__gte=cutoff_dob, date_of_birth__isnull=False)
            elif age_operator == 'lte':
                qs = qs.filter(date_of_birth__gte=cutoff_dob, date_of_birth__isnull=False)
            elif age_operator == 'eq':
                try:
                    lower_dob = date(today.year - aval - 1, today.month, today.day)
                except ValueError:
                    lower_dob = date(today.year - aval - 1, today.month, 28)
                upper_dob = cutoff_dob
                qs = qs.filter(date_of_birth__gt=lower_dob, date_of_birth__lte=upper_dob, date_of_birth__isnull=False)
        except (ValueError, TypeError):
            pass

    # Inquiry status filter
    if inquiry_status:
        from inquiries.models import Inquiry
        if inquiry_status == 'has_open':
            qs = qs.filter(submitted_inquiries__status__in=['open', 'in_progress', 'pending_info']).distinct()
        elif inquiry_status == 'no_inquiries':
            qs = qs.filter(submitted_inquiries__isnull=True)
        elif inquiry_status == 'has_pending':
            qs = qs.filter(submitted_inquiries__status='pending_info').distinct()
        elif inquiry_status == 'resolved':
            qs = qs.filter(submitted_inquiries__status='resolved').distinct()
        elif inquiry_status == 'has_any':
            qs = qs.filter(submitted_inquiries__isnull=False).distinct()

    return qs


@admin_required
def export_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    report_type = request.GET.get('report_type', 'employee_list')
    employees = get_filtered_employees(request)

    wb = Workbook()
    ws = wb.active

    # Header style
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1B4F8A', end_color='1B4F8A', fill_type='solid')

    if report_type == 'employee_list':
        ws.title = 'Employee List'
        headers = ['#', 'Employee No.', 'Full Name', 'Email', 'Entity Type', 'Entity',
                   'Cadre Category', 'Speciality', 'Position', 'Employee Type', 'Date Joined', 'Profile %']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for row_num, emp in enumerate(employees, 2):
            ws.cell(row=row_num, column=1, value=row_num - 1)
            ws.cell(row=row_num, column=2, value=emp.employee_number)
            ws.cell(row=row_num, column=3, value=emp.user.get_full_name())
            ws.cell(row=row_num, column=4, value=emp.user.email)
            ws.cell(row=row_num, column=5, value=emp.get_entity_type_display() if emp.entity_type else '')
            ws.cell(row=row_num, column=6, value=emp.get_entity_name())
            ws.cell(row=row_num, column=7, value=emp.cadre_category.name if emp.cadre_category else '')
            ws.cell(row=row_num, column=8, value=emp.position.name if emp.position else '')
            ws.cell(row=row_num, column=9, value=emp.job_rank.name if emp.job_rank else '')
            ws.cell(row=row_num, column=10, value=emp.employee_type.name if emp.employee_type else '')
            ws.cell(row=row_num, column=11, value=str(emp.date_joined_ministry) if emp.date_joined_ministry else '')
            ws.cell(row=row_num, column=12, value=f"{emp.profile_completion}%")

        # Auto-width
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

    elif report_type == 'contract_expiry':
        ws.title = 'Contract Expiry'
        employees = employees.filter(contract_end_date__isnull=False).order_by('contract_end_date')
        headers = ['#', 'Employee No.', 'Full Name', 'Entity', 'Speciality', 'Contract End Date', 'Days Remaining']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        today = timezone.now().date()
        for row_num, emp in enumerate(employees, 2):
            days = (emp.contract_end_date - today).days if emp.contract_end_date else None
            ws.cell(row=row_num, column=1, value=row_num - 1)
            ws.cell(row=row_num, column=2, value=emp.employee_number)
            ws.cell(row=row_num, column=3, value=emp.user.get_full_name())
            ws.cell(row=row_num, column=4, value=emp.get_entity_name())
            ws.cell(row=row_num, column=5, value=emp.position.name if emp.position else '')
            ws.cell(row=row_num, column=6, value=str(emp.contract_end_date))
            ws.cell(row=row_num, column=7, value=days)

    elif report_type == 'time_in_position':
        ws.title = 'Time in Speciality'
        employees_list = employees.filter(date_joined_position__isnull=False).order_by('date_joined_position')
        headers = ['#', 'Employee No.', 'Full Name', 'Entity', 'Cadre Category', 'Speciality', 'Date Joined Speciality', 'Years in Speciality']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        today = timezone.now().date()
        for row_num, emp in enumerate(employees_list, 2):
            days = (today - emp.date_joined_position).days if emp.date_joined_position else 0
            years = round(days / 365.25, 1)
            ws.cell(row=row_num, column=1, value=row_num - 1)
            ws.cell(row=row_num, column=2, value=emp.employee_number)
            ws.cell(row=row_num, column=3, value=emp.user.get_full_name())
            ws.cell(row=row_num, column=4, value=emp.get_entity_name())
            ws.cell(row=row_num, column=5, value=emp.cadre_category.name if emp.cadre_category else '')
            ws.cell(row=row_num, column=6, value=emp.position.name if emp.position else '')
            ws.cell(row=row_num, column=7, value=str(emp.date_joined_position))
            ws.cell(row=row_num, column=8, value=years)
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

    elif report_type == 'deployment_by_entity':
        ws.title = 'Deployment by Entity'
        headers = ['#', 'Employee No.', 'Full Name', 'Entity Type', 'Entity', 'Cadre Category', 'Speciality', 'Position', 'Date Joined']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        for row_num, emp in enumerate(employees, 2):
            ws.cell(row=row_num, column=1, value=row_num - 1)
            ws.cell(row=row_num, column=2, value=emp.employee_number)
            ws.cell(row=row_num, column=3, value=emp.user.get_full_name())
            ws.cell(row=row_num, column=4, value=emp.get_entity_type_display() if emp.entity_type else '')
            ws.cell(row=row_num, column=5, value=emp.get_entity_name())
            ws.cell(row=row_num, column=6, value=emp.cadre_category.name if emp.cadre_category else '')
            ws.cell(row=row_num, column=7, value=emp.position.name if emp.position else '')
            ws.cell(row=row_num, column=8, value=emp.job_rank.name if emp.job_rank else '')
            ws.cell(row=row_num, column=9, value=str(emp.date_joined_ministry) if emp.date_joined_ministry else '')
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

    # Title row
    ws.insert_rows(1)
    title_cell = ws.cell(row=1, column=1, value=f"IT Cadre Tracking Database - {ws.title} Report")
    title_cell.font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers) if 'headers' in dir() else 12)
    ws.insert_rows(2)
    ws.cell(row=2, column=1, value=f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"cadre_report_{report_type}_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@admin_required
def export_pdf(request):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors

    report_type = request.GET.get('report_type', 'employee_list')
    employees = get_filtered_employees(request)

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=1 * cm, bottomMargin=1 * cm)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=16, textColor=colors.HexColor('#1B4F8A'))
    elements.append(Paragraph("IT Cadre Tracking Database", title_style))
    elements.append(Paragraph("Ministry of ICT and National Guidance, Uganda", styles['Normal']))
    elements.append(Spacer(1, 0.3 * cm))
    elements.append(Paragraph(f"Report: Employee List | Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}",
                               styles['Normal']))
    elements.append(Spacer(1, 0.5 * cm))

    # Table
    data = [['#', 'Emp. No.', 'Full Name', 'Entity Type', 'Entity', 'Cadre Category', 'Position', 'Profile %']]
    for i, emp in enumerate(employees[:200], 1):
        data.append([
            str(i),
            emp.employee_number,
            emp.user.get_full_name(),
            emp.get_entity_type_display() if emp.entity_type else '',
            emp.get_entity_name(),
            emp.cadre_category.name if emp.cadre_category else '',
            emp.position.name if emp.position else '',
            f"{emp.profile_completion}%",
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B4F8A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f7fa')]),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)

    doc.build(elements)
    output.seek(0)
    filename = f"cadre_report_{timezone.now().strftime('%Y%m%d')}.pdf"
    response = HttpResponse(output.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@admin_required
def chart_data(request):
    """API endpoint for chart data based on current filters."""
    employees = get_filtered_employees(request)

    # Distribution by cadre category
    cadre = list(employees.values('cadre_category__name').annotate(count=Count('id')).order_by('-count'))
    for c in cadre:
        c['cadre_category__name'] = c['cadre_category__name'] or 'Unassigned'

    # Distribution by entity type
    entity = list(employees.values('entity_type').annotate(count=Count('id')).order_by('-count'))
    entity_map = dict(Employee._meta.get_field('entity_type').choices)
    for e in entity:
        e['label'] = entity_map.get(e['entity_type'], e['entity_type'] or 'Unknown')

    # Distribution by gender
    gender = list(employees.values('gender').annotate(count=Count('id')).order_by('-count'))
    for g in gender:
        g['label'] = {'M': 'Male', 'F': 'Female', '': 'Not Set'}.get(g['gender'], g['gender'] or 'Not Set')

    # Distribution by position (top 10)
    position = list(employees.values('position__name').annotate(count=Count('id')).order_by('-count')[:10])
    for p in position:
        p['position__name'] = p['position__name'] or 'Unassigned'

    return JsonResponse({
        'cadre': cadre,
        'entity': entity,
        'gender': gender,
        'position': position,
        'total': employees.count(),
    })
